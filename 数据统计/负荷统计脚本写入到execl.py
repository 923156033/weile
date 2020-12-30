#-*- codeing = utf-8 -*-
#@Time : 2020/11/28  14:01
#@Author : cyz
#@FIle : 负荷脚本.py
#@Software: PyCharm
import pandas as pd
import xlwt
import os
from openpyxl.reader.excel import load_workbook
#定义类，查询地区负荷，求出最大、最小、平均值、求和
def handle_max(diqu):
    data1=data[data['服务器名称'].str.contains(diqu)]
    max1 = data1['负荷'].max()
    return [max1]
def handle_min(diqu):
    data1=data[data['服务器名称'].str.contains(diqu)]
    min1 = data1['负荷'].min()
    return [min1]
def handle_mean(diqu):
    data1=data[data['服务器名称'].str.contains(diqu)]
    mean1 = int(data1['负荷'].mean())
    return [mean1]
def handle_sum(diqu):
    data1=data[data['服务器名称'].str.contains(diqu)]
    sum1 = data1['负荷'].sum()
    return [sum1]
area=["通用","四川_甘肃_宁夏_云南","陕西","山西_内蒙","山东","辽宁","江西_福建","江苏_安徽_浙江_上海","吉林","湖南","湖北","黑龙江","河南","河北_北京_天津","贵州","广东_广西_海南","高防" ]
data = pd.read_csv(r"E:\GitHub\weile\数据统计\导出数据\地区导出数据 (1).csv",encoding='utf-8')
workbook = xlwt.Workbook(encoding="utf-8")
file_home = r'E:\GitHub\weile\数据统计\统计负荷模板.xlsx'
wb = load_workbook(filename=file_home)
ws = wb['Sheet1']
a=b=c=d=0
#微乐小程序代理最大值、最小值、平均值、总负荷写入exel
for i in range(4, 21):
    max1 = handle_max(area[a])
    ws['H' + str(i)] = max1[0]
    a=a+1
for i in range(4, 21):
    min1 = handle_min(area[b])
    ws['I' + str(i)] = min1[0]
    b=b+1
for i in range(4, 21):
    mean1 = handle_mean(area[c])
    ws['J' + str(i)] = mean1[0]
    c=c+1
for i in range(4, 21):
    sum1 = handle_sum(area[d])
    ws['K' + str(i)] = sum1[0]
    d=d+1
#微乐小程序大厅最大值、最小值、平均值、总负荷写入exel
data = pd.read_csv(r"E:\GitHub\weile\数据统计\导出数据\地区导出数据 (2).csv",encoding='utf-8')
a=b=c=d=0
for i in range(4, 21):
    max1 = handle_max(area[a])
    ws['N' + str(i)] = max1[0]
    a=a+1
for i in range(4, 21):
    min1 = handle_min(area[b])
    ws['O' + str(i)] = min1[0]
    b=b+1
for i in range(4, 21):
    mean1 = handle_mean(area[c])
    ws['P' + str(i)] = mean1[0]
    c=c+1
for i in range(4, 21):
    sum1 = handle_sum(area[d])
    ws['R' + str(i)] = sum1[0]
    d=d+1
#微乐APP大厅最大值、最小值、平均值、总负荷写入exel
data = pd.read_csv(r"E:\GitHub\weile\数据统计\导出数据\地区导出数据.csv",encoding='utf-8')
max1 = handle_max(area[0])
ws['b4'] = max1[0]
min1 = handle_min(area[0])
ws['C4'] = min1[0]
mean1 = handle_mean(area[0])
ws['D4'] = mean1[0]
sum1 = handle_sum(area[0])
ws['E4'] = sum1[0]
wb.save(file_home)
#吉祥APP大厅最大值、最小值、平均值、总负荷写入exel
area=["大厅"]
data = pd.read_csv(r"E:\GitHub\weile\数据统计\导出数据\地区导出数据 (4).csv",encoding='utf-8')
max1 = handle_max(area[0])
ws['b12'] = max1[0]
min1 = handle_min(area[0])
ws['C12'] = min1[0]
mean1 = handle_mean(area[0])
ws['D12'] = mean1[0]
sum1 = handle_sum(area[0])
ws['E12'] = sum1[0]
wb.save(file_home)
#吉祥小程序代理最大值、最小值、平均值、总负荷写入exel
area=["小程序[2-9]+--"]
data = pd.read_csv(r"E:\GitHub\weile\数据统计\导出数据\地区导出数据 (5).csv",encoding='utf-8')
max1 = handle_max(area[0])
ws['b20'] = max1[0]
min1 = handle_min(area[0])
ws['C20'] = min1[0]
mean1 = handle_mean(area[0])
ws['D20'] = mean1[0]
sum1 = handle_sum(area[0])
ws['E20'] = sum1[0]
wb.save(file_home)
#吉祥小程序大厅最大值、最小值、平均值、总负荷写入exel
area=["小程序"]
data = pd.read_csv(r"E:\GitHub\weile\数据统计\导出数据\地区导出数据 (6).csv",encoding='utf-8')
max1 = handle_max(area[0])
ws['b16'] = max1[0]
min1 = handle_min(area[0])
ws['C16'] = min1[0]
mean1 = handle_mean(area[0])
ws['D16'] = mean1[0]
sum1 = handle_sum(area[0])
ws['E16'] = sum1[0]
#微乐用户负荷数
area=["[\S]"]
data = pd.read_csv(r"E:\GitHub\weile\数据统计\导出数据\地区导出数据 (3).csv",encoding='utf-8')
sum1 = handle_sum(area[0])
ws['D5'] = sum1[0]
#吉祥用户负荷数
data = pd.read_csv(r"E:\GitHub\weile\数据统计\导出数据\地区导出数据 (7).csv",encoding='utf-8')
sum1 = handle_sum(area[0])
ws['D7'] = sum1[0]
wb.save(file_home)