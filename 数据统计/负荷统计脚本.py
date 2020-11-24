# -*- coding: utf-8 -*-
import pandas as pd
import re
import os.path
import os
#每次使用这个程序之前先在中班统计负荷文件夹下复制一个"微乐吉祥中班统计负荷模板新xxxx.xlsx"
#导出顺序有要求jx 大厅 代理 小程序，然后微乐大厅、代理、小程序，而且google下载文件夹下必须没有之前导出的文件，需要清空之前导出的以地区导出数据名字开头的文件
from openpyxl.reader.excel import load_workbook
data = pd.read_csv(r'E:\GitHub\weile\数据统计\导出数据\地区导出数据.csv',encoding='utf-8')
data1 = pd.read_csv(r'E:\GitHub\weile\数据统计\导出数据\地区导出数据 (1).csv',encoding='utf-8')
pattern = re.compile('厅[0-9]+--')
pattern1 = re.compile('序[0-9]+--')
data.replace(pattern,"厅",inplace=True)
data1.replace(pattern1,"序",inplace=True)
data_new = data.drop([0]) # 删除第一行数据
data1_new = data1.drop([0,1])  # 删除0,1行数据
#print(data_new)
#print(data1_new)
jxapp_max=int(data_new.groupby('服务器名称')['负荷'].max())
jxapp_min=int(data_new.groupby('服务器名称')['负荷'].min())
jxapp_avg=int(data_new.groupby('服务器名称')['负荷'].mean())
jxxiaochengxudaili_max=int(data1_new.groupby('服务器名称')['负荷'].max())
jxxiaochengxudaili_min=int(data1_new.groupby('服务器名称')['负荷'].min())
jxxiaochengxudaili_avg=int(data1_new.groupby('服务器名称')['负荷'].mean())
#print(jxapp_max,jxapp_min,jxapp_avg)
#print(jxxiaochengxudaili_max,jxxiaochengxudaili_min,jxxiaochengxudaili_avg)
# 、、、、对于data进行多次操作，如果想要连续操作，记得都将.号之前的主语改成同一pandas对象，
# 比如前来两个操作，第二个主语需要改成data_new对象。如果想要保存新的csv文件，则为：


#吉祥小程序
data2 = pd.read_csv(r'E:\GitHub\weile\数据统计\导出数据\地区导出数据 (2).csv',encoding='utf-8')
pattern2 = re.compile('序[0-9]+--')
data2.replace(pattern1,"序",inplace=True)
#print(data2)
jxxiaochengxu_max=int(data2.groupby('服务器名称')['负荷'].max())
jxxiaochengxu_min=int(data2.groupby('服务器名称')['负荷'].min())
jxxiaochengxu_avg=int(data2.groupby('服务器名称')['负荷'].mean())
# print(jxxiaochengxu_max,jxxiaochengxu_min,jxxiaochengxu_avg)


#微乐APP
data3 = pd.read_csv(r'E:\GitHub\weile\数据统计\导出数据\地区导出数据 (3).csv',encoding='utf-8')
#data1_new.to_csv(r'E:\GitHub\weile\数据统计\导出数据\吉祥小程序.csv',encoding="utf_8_sig",index=0)
pattern3 = re.compile('_[0-9]+')
data3.replace(pattern3,"",inplace=True)
data3_new = data3.drop([0]) # 删除第一行数据
#print(data3_new)
wlapp_max=int(data3_new.groupby('服务器名称')['负荷'].max())
wlapp_min=int(data3_new.groupby('服务器名称')['负荷'].min())
wlapp_avg=int(data3_new.groupby('服务器名称')['负荷'].mean())
wlapp_sum=int(data3_new.groupby('服务器名称')['负荷'].sum())
# print(wlapp_max,wlapp_min,wlapp_avg)

#微乐代理
data4 = pd.read_csv(r'E:\GitHub\weile\数据统计\导出数据\地区导出数据 (4).csv',encoding='utf-8')
pattern4 = re.compile('_[0-9]+')
data4.replace(pattern4,"",inplace=True)
data4_new=data4.sort_values(by='服务器名称',ascending=False)
# print(data4_new)
wldaili_max=list(data4_new.groupby('服务器名称')['负荷'].max())
wldaili_min=list(data4_new.groupby('服务器名称')['负荷'].min())
wldaili_avg=list(data4_new.groupby('服务器名称')['负荷'].mean())
wldaili_sum=list(data4_new.groupby('服务器名称')['负荷'].sum())
wldaili_avg=list(map(int,wldaili_avg))
#print(wldaili_max,wldaili_min,wldaili_avg)
#data4_new.to_csv(r'E:\GitHub\weile\数据统计\导出数据\微乐代理.csv',encoding="utf_8_sig",index=0)
wldailimaxlist=[wldaili_max[13],wldaili_max[1],wldaili_max[14],wldaili_max[3],wldaili_max[2],wldaili_max[12],wldaili_max[6],wldaili_max[5],wldaili_max[0],wldaili_max[10],wldaili_max[9],wldaili_max[16],wldaili_max[8],wldaili_max[7],wldaili_max[11],wldaili_max[4],wldaili_max[15]]
wldailiminlist=[wldaili_min[13],wldaili_min[1],wldaili_min[14],wldaili_min[3],wldaili_min[2],wldaili_min[12],wldaili_min[6],wldaili_min[5],wldaili_min[0],wldaili_min[10],wldaili_min[9],wldaili_min[16],wldaili_min[8],wldaili_min[7],wldaili_min[11],wldaili_min[4],wldaili_min[15]]
wldailiavglist=[wldaili_avg[13],wldaili_avg[1],wldaili_avg[14],wldaili_avg[3],wldaili_avg[2],wldaili_avg[12],wldaili_avg[6],wldaili_avg[5],wldaili_avg[0],wldaili_avg[10],wldaili_avg[9],wldaili_avg[16],wldaili_avg[8],wldaili_avg[7],wldaili_avg[11],wldaili_avg[4],wldaili_avg[15]]
wldailisumlist=[wldaili_sum[13],wldaili_sum[1],wldaili_sum[14],wldaili_sum[3],wldaili_sum[2],wldaili_sum[12],wldaili_sum[6],wldaili_sum[5],wldaili_sum[0],wldaili_sum[10],wldaili_sum[9],wldaili_sum[16],wldaili_sum[8],wldaili_sum[7],wldaili_sum[11],wldaili_sum[4],wldaili_sum[15]]
# print(wldailimaxlist)
# print(wldailiminlist)
# print(wldailiavglist)


#微乐小程序
data5 = pd.read_csv(r'E:\GitHub\weile\数据统计\导出数据\地区导出数据 (5).csv',encoding='utf-8')
pattern5 = re.compile('_[0-9]+')
data5.replace(pattern5,"",inplace=True)
data5_new=data5.sort_values(by='服务器名称',ascending=False)
# print(data5_new)
wlxiaochengxu_max=list(data5_new.groupby('服务器名称')['负荷'].max())
wlxiaochengxu_min=list(data5_new.groupby('服务器名称')['负荷'].min())
wlxiaochengxu_avg=list(data5_new.groupby('服务器名称')['负荷'].mean())
wlxiaochengxu_sum=list(data5_new.groupby('服务器名称')['负荷'].sum())
wlxiaochengxu_avg=list(map(int,wlxiaochengxu_avg))
wlxiaochengxumaxlist=[wlxiaochengxu_max[13],wlxiaochengxu_max[1],wlxiaochengxu_max[14],wlxiaochengxu_max[3],wlxiaochengxu_max[2],wlxiaochengxu_max[12],wlxiaochengxu_max[6],wlxiaochengxu_max[5],wlxiaochengxu_max[0],wlxiaochengxu_max[10],wlxiaochengxu_max[9],wlxiaochengxu_max[16],wlxiaochengxu_max[8],wlxiaochengxu_max[7],wlxiaochengxu_max[11],wlxiaochengxu_max[4],wlxiaochengxu_max[15]]
wlxiaochengxuminlist=[wlxiaochengxu_min[13],wlxiaochengxu_min[1],wlxiaochengxu_min[14],wlxiaochengxu_min[3],wlxiaochengxu_min[2],wlxiaochengxu_min[12],wlxiaochengxu_min[6],wlxiaochengxu_min[5],wlxiaochengxu_min[0],wlxiaochengxu_min[10],wlxiaochengxu_min[9],wlxiaochengxu_min[16],wlxiaochengxu_min[8],wlxiaochengxu_min[7],wlxiaochengxu_min[11],wlxiaochengxu_min[4],wlxiaochengxu_min[15]]
wlxiaochengxuavglist=[wlxiaochengxu_avg[13],wlxiaochengxu_avg[1],wlxiaochengxu_avg[14],wlxiaochengxu_avg[3],wlxiaochengxu_avg[2],wlxiaochengxu_avg[12],wlxiaochengxu_avg[6],wlxiaochengxu_avg[5],wlxiaochengxu_avg[0],wlxiaochengxu_avg[10],wlxiaochengxu_avg[9],wlxiaochengxu_avg[16],wlxiaochengxu_avg[8],wlxiaochengxu_avg[7],wlxiaochengxu_avg[11],wlxiaochengxu_avg[4],wlxiaochengxu_avg[15]]
wlxiaochengxusumlist=[wlxiaochengxu_sum[13],wlxiaochengxu_sum[1],wlxiaochengxu_sum[14],wlxiaochengxu_sum[3],wlxiaochengxu_sum[2],wlxiaochengxu_sum[12],wlxiaochengxu_sum[6],wlxiaochengxu_sum[5],wlxiaochengxu_sum[0],wlxiaochengxu_sum[10],wlxiaochengxu_sum[9],wlxiaochengxu_sum[16],wlxiaochengxu_sum[8],wlxiaochengxu_sum[7],wlxiaochengxu_sum[11],wlxiaochengxu_sum[4],wlxiaochengxu_sum[15]]
# print(wlxiaochengxumaxlist)
# print(wlxiaochengxuminlist)
# print(wlxiaochengxuavglist)


#数据输出到表格里
file_home = r'E:\GitHub\weile\数据统计\统计负荷模板.xlsx'
wb = load_workbook(filename=file_home)  # 打开excel文件
sheet_ranges = wb['Sheet1']
#print(sheet_ranges['A1'].value)  # 打印A1单元格的值
ws = wb['Sheet1']  # 根据Sheet1这个sheet名字来获取该sheet
# 吉祥APP大厅
ws['b12'] = jxapp_max
ws['c12'] = jxapp_min
ws['d12'] = jxapp_avg
#吉祥小程序大厅
ws['b16'] = jxxiaochengxu_max
ws['c16'] = jxxiaochengxu_min
ws['d16'] = jxxiaochengxu_avg
#吉祥小程序代理
ws['b20'] = jxxiaochengxudaili_max
ws['c20'] = jxxiaochengxudaili_min
ws['d20'] = jxxiaochengxudaili_avg

#微乐app大厅
ws['b4'] = wlapp_max
ws['c4'] = wlapp_min
ws['d4'] = wlapp_avg
k=0
#微乐代理
for i in range(4, 21):
            ws['g'+str(i)]= wldailimaxlist[k]
            ws['h' + str(i)] = wldailiminlist[k]
            ws['i' + str(i)] = wldailiavglist[k]
            k=k+1
#微乐小程序
k=0
for i in range(4, 21):
            ws['l'+str(i)]= wlxiaochengxumaxlist[k]
            ws['m' + str(i)] = wlxiaochengxuminlist[k]
            ws['n' + str(i)] = wlxiaochengxuavglist[k]
            k=k+1
wb.save(file_home)  # 保存修改后的excel
print("微乐小程序求和")
#小程序求和
j=0
while j<len(wlxiaochengxu_sum):
    print(wlxiaochengxusumlist[j])
    j+=1
#代理求和
print("微乐代理求和")
i=0
while i<len(wldailisumlist):
    print(wldailisumlist[i])
    i+=1
#微乐app求和
print("微乐app求和")
print(wlapp_sum)