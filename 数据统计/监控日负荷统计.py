#!C:\Program Files\Python38
# -*- coding: utf-8 -*-

from aliyun.log import LogClient,GetLogsRequest
from openpyxl.styles import Font,colors,Alignment,PatternFill,Border,Side  #设定字体，字体颜色，文字位置，背景色，框线，框线样式
import openpyxl
import datetime, time
import os

#微乐日志服务api
wlclient = LogClient()
wlproject = '' #projectname
wlusername = 'X_用户服务器'
wlhalls = ['小游戏大厅_通用', '小游戏大厅_四川_甘肃_宁夏_云南', '小游戏大厅_陕西', '小游戏大厅_山西_内蒙', '小游戏大厅_山东', '小游戏大厅_辽宁', '小游戏大厅_江西_福建',
                  '小游戏大厅_江苏_安徽_浙江_上海', '小游戏大厅_吉林', '小游戏大厅_湖南', '小游戏大厅_湖北', '小游戏大厅_黑龙江', '小游戏大厅_河南', '小游戏大厅_河北_北京_天津',
                  '小游戏大厅_贵州', '小游戏大厅_广东_广西_海南', '小游戏大厅_高防', 'APP大厅_通用']
#吉祥日志服务api
jxclient = LogClient()
jxproject = ''  #projectname
jxusername = '用户服务器'
jxhalls = ['吉祥大厅', '小程序']


class create_xls_data:
    def __init__(self,client,project,username,halls):
        self.client = client
        self.project = project
        self.username = username
        self.halls = halls
        self.fromtime = int(time.time()) - 7200
        self.totime = int(time.time())
        self.get_log_data()

    def get_log_data(self):
        request = GetLogsRequest()
        request.set_project(self.project)
        request.set_logstore('server')
        request.set_from(self.fromtime)
        request.set_to(self.totime)
        request.set_query('name = {}'.format(self.username))
        # 获取两小时内用户服务器最大时间戳
        self.userlogs = self.client.get_logs(request).body
        try:
            maxus = self.get_max_time(self.userlogs)
            self.max_time = maxus.get('time')
            print('2小时内最高负荷取值时间为：{}，{}负荷数为：{}'.format(self.max_time, self.username, maxus.get('connection')))
        except:
            self.max_time = self.userlogs[-1].get('time')
            print('高峰期时间取值异常,使用最近一次{}取值时间'.format(self.username))
            print('负荷取值时间为：{}，{}负荷数为：{}'.format(self.max_time, self.username, self.userlogs[-1].get('connection')))
        max_timeStamp = time.mktime(time.strptime(self.max_time, "%Y-%m-%d %H:%M:%S"))
        # 根据最大时间戳获取日志
        request.set_from(max_timeStamp)
        request.set_to(max_timeStamp + 1)
        request.set_query('')
        self.logs_list, l = [], 0
        while True:
            logs = self.client.get_logs(request)
            self.logs_list.extend(logs.get_body())
            if int(logs.headers['x-log-count']) != 100:
                break
            l += 100
            request.set_offset(l)
        #写入日志数据
        self.insert_log_data()

    def get_max_time(self,__us):
        for i in range(1, len(__us)):
            if str(__us[i - 1].get('connection')) >= str(__us[i].get('connection')):
                __us[i - 1], __us[i] = __us[i], __us[i - 1]
            maxus = __us[i]
        return maxus

    def insert_log_data(self):
        hallsum,hallnum,proxysum,proxynum,num = 0,0,0,0,4
        for hall in self.halls:
            hallcon,proxycon = [],[]
            for logs in self.logs_list:
                if logs.get('region_name').startswith(hall):
                    if '反向代理' in logs.get('name'):
                        if int(logs.get('connection')) <= 10 and '高防' not in logs.get('region_name'):
                            print('"{}"的"{}"负荷值为：{}，低于正常值,不记录'.format(logs.get('region_name'),logs.get('name'),logs.get('connection')))
                        else:
                            proxycon.append(int(logs.get('connection')))
                    elif 'APP' in logs.get('name') or '小游戏' in logs.get('name') or logs.get('name').startswith('大厅服务器') or '高防' in logs.get('name'):
                        if int(logs.get('connection')) <= 10 and '高防' not in logs.get('region_name'):
                            print(' {} 的 {} 负荷值为：{}，低于正常值,不记录'.format(logs.get('region_name'),logs.get('name'),logs.get('connection')))
                        else:
                            hallcon.append(int(logs.get('connection')))
            if hallcon or proxycon:
                if hall.startswith('小游戏大厅'):
                    ws['F{}'.format(num)] = '{} ({})'.format(hall,len(proxycon))
                    ws['G{}'.format(num)] = max(proxycon)
                    ws['H{}'.format(num)] = min(proxycon)
                    ws['I{}'.format(num)] = sum(proxycon) // len(proxycon)
                    ws['K{}'.format(num)] = '{} ({})'.format(hall, len(hallcon))
                    ws['L{}'.format(num)] = max(hallcon)
                    ws['M{}'.format(num)] = min(hallcon)
                    ws['N{}'.format(num)] = sum(hallcon) // len(hallcon)
                    ws2['A{}'.format(num + 2)] = hall
                    ws2['B{}'.format(num + 2)] = sum(hallcon)
                    ws2['C{}'.format(num + 2)] = len(hallcon)
                    ws2['A{}'.format(num + 22)] = hall
                    ws2['B{}'.format(num + 22)] = sum(proxycon)
                    ws2['C{}'.format(num + 22)] = len(proxycon)
                    hallsum += sum(hallcon)
                    hallnum += len(hallcon)
                    proxysum += sum(proxycon)
                    proxynum += len(proxycon)
                elif hall.startswith('APP大厅_通用'):
                    ws['A{}'.format(4)] = '{} ({})'.format(hall, len(hallcon))
                    ws['B{}'.format(4)] = max(hallcon)
                    ws['C{}'.format(4)] = min(hallcon)
                    ws['D{}'.format(4)] = sum(hallcon) // len(hallcon)
                    ws2['A3'] = hall
                    ws2['B3'] = sum(hallcon)
                    ws2['C3'] = len(hallcon)
                elif hall.startswith('吉祥大厅'):
                    ws['A{}'.format(num+8)] = '{} ({})'.format(hall, len(hallcon))
                    ws['B{}'.format(num+8)] = max(hallcon)
                    ws['C{}'.format(num+8)] = min(hallcon)
                    ws['D{}'.format(num+8)] = sum(hallcon) // len(hallcon)
                    ws2['A46'] = hall
                    ws2['B46'] = sum(hallcon)
                    ws2['C46'] = len(hallcon)
                elif hall.startswith('小程序'):
                    ws['A{}'.format(16)] = '{} ({})'.format(hall, len(proxycon))
                    ws['B{}'.format(16)] = max(proxycon)
                    ws['C{}'.format(16)] = min(proxycon)
                    ws['D{}'.format(16)] = sum(proxycon) // len(proxycon)
                    ws['A{}'.format(20)] = '{} ({})'.format(hall, len(hallcon))
                    ws['B{}'.format(20)] = max(hallcon)
                    ws['C{}'.format(20)] = min(hallcon)
                    ws['D{}'.format(20)] = sum(hallcon) // len(hallcon)
                    ws2['A49'] = hall
                    ws2['B49'] = sum(hallcon)
                    ws2['C49'] = len(hallcon)
                    ws2['A52'] = hall
                    ws2['B52'] = sum(proxycon)
                    ws2['C52'] = len(proxycon)
            else:
                print('{}取值异常'.format(hall))
            num += 1
        if self.username == 'X_用户服务器':
            ws2['A23'] = '总计'
            ws2['B23'] = hallsum
            ws2['C23'] = hallnum
            ws2['A43'] = '总计'
            ws2['B43'] = proxysum
            ws2['C43'] = proxynum

def Insert_Template(wltime,jxtime):
    title_font = Font(name='宋体', bold=True, size=16)
    table_font = Font(name='宋体', bold=True, size=11)
    tables = ws['A1':'N20']
    for tablecell in tables:
        for tablerow in tablecell:
            tablerow.alignment = Alignment(vertical='center')
    for m in range(1, 21):
        ws.row_dimensions[m].height = 21.6
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['F'].width = 34
    ws.column_dimensions['K'].width = 34
    ws.column_dimensions['E'].width = 4
    ws.column_dimensions['J'].width = 4
    ws2.column_dimensions['A'].width = 34
    titles = ['A1', 'F1', 'K1', 'A9', 'A13', 'A17']
    for titlerow in titles:
        ws[titlerow].font = title_font
        ws[titlerow].alignment = Alignment(horizontal='center', vertical='center')
        if titlerow == 'A1':
            ws.merge_cells('A1:D2')
            ws[titlerow] = '微乐APP大厅 {}'.format(wltime)
        elif titlerow == 'F1':
            ws.merge_cells('F1:I2')
            ws[titlerow] = '微乐小程序代理 {}'.format(wltime)
        elif titlerow == 'K1':
            ws.merge_cells('K1:N2')
            ws[titlerow] = '微乐小程序大厅 {}'.format(wltime)
        elif titlerow == 'A9':
            ws.merge_cells('A9:D10')
            ws[titlerow] = '吉祥APP大厅 {}'.format(jxtime)
        elif titlerow == 'A13':
            ws.merge_cells('A13:D14')
            ws[titlerow] = '吉祥小程序代理 {}'.format(jxtime)
        elif titlerow == 'A17':
            ws.merge_cells('A17:D18')
            ws[titlerow] = '吉祥小程序大厅 {}'.format(jxtime)
    titles2 = ['A1', 'A4', 'A24', 'A44', 'A47', 'A50']
    for title2row in titles2:
        ws2[title2row].font = title_font
        ws2[title2row].alignment = Alignment(horizontal='center', vertical='center')
        if title2row == 'A1':
            ws2.merge_cells('A1:C1')
            ws2[title2row] = '微乐APP大厅 {}'.format(wltime)
        elif title2row == 'A4':
            ws2.merge_cells('A4:C4')
            ws2[title2row] = '微乐小程序大厅{}'.format(wltime)
        elif title2row == 'A24':
            ws2.merge_cells('A24:C24')
            ws2[title2row] = '微乐小程序代理 {}'.format(wltime)
        elif title2row == 'A44':
            ws2.merge_cells('A44:C44')
            ws2[title2row] = '吉祥APP大厅 {}'.format(jxtime)
        elif title2row == 'A47':
            ws2.merge_cells('A47:C47')
            ws2[title2row] = '吉祥小程序大厅 {}'.format(jxtime)
        elif title2row == 'A50':
            ws2.merge_cells('A50:C50')
            ws2[title2row] = '吉祥小程序反代 {}'.format(jxtime)
    tabname = ['A3', 'F3', 'K3', 'A11', 'A15', 'A19', 'B3', 'G3', 'L3', 'B11', 'B15', 'B19', 'C3', 'H3', 'M3', 'C11',
               'C15', 'C19', 'D3', 'I3', 'N3', 'D11', 'D15', 'D19']
    for row in tabname:
        ws[row].font = table_font
        if row.startswith('A') or row.startswith('F') or row.startswith('K'):
            ws[row] = '服务器组'
        elif row.startswith('B') or row.startswith('G') or row.startswith('L'):
            ws[row] = '最高负荷'
        elif row.startswith('C') or row.startswith('H') or row.startswith('M'):
            ws[row] = '最低负荷'
        elif row.startswith('D') or row.startswith('I') or row.startswith('N'):
            ws[row] = '平均负荷'
    tab2name = ['A2', 'B2', 'C2', 'A5', 'B5', 'C5', 'A25', 'B25', 'C25', 'A45', 'B45', 'C45', 'A48', 'B48', 'C48',
                'A51', 'B51', 'C51']
    for row2 in tab2name:
        ws2[row2].font = table_font
        if row2.startswith('A'):
            ws2[row2] = '服务器名称'
        elif row2.startswith('B'):
            ws2[row2] = '总负荷'
        elif row2.startswith('C'):
            ws2[row2] = '服务数量'
    wl1, wl2, wl3 = ws['A1':'D4'], ws['F1':'I20'], ws['K1':'N20']
    jx1 = ws['A9':'D20']
    thin = Side(border_style='thin', color='000000')
    double = Side(border_style='double', color='ff0000')
    for wlrange in wl1, wl2, wl3:
        for wlcell in wlrange:
            for wlrow in wlcell:
                wlrow.fill = PatternFill(patternType='solid', fgColor='D8E4BC')
                wlrow.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for jxcell in jx1:
        for jxrow in jxcell:
            jxrow.fill = PatternFill(patternType='solid', fgColor='DCE6F1')
            jxrow.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    conn = ws2['A1':'C52']
    for cell2 in conn:
        for connrow2 in cell2:
            connrow2.border = Border(top=thin, left=thin, right=thin, bottom=thin)

if __name__ == '__main__':
    wb = openpyxl.Workbook()
    ws = wb.active
    ws2 = wb.create_sheet(title='统计')
    weile = create_xls_data(wlclient,wlproject,wlusername,wlhalls)
    jixiang = create_xls_data(jxclient,jxproject,jxusername,jxhalls)
    Insert_Template(weile.max_time,jixiang.max_time)
    try:
        wb.save('负荷统计{}.xlsx'.format(datetime.datetime.now().strftime("%Y-%m-%d")))
        print('已输出文件：负荷统计{}.xlsx'.format(datetime.datetime.now().strftime("%Y-%m-%d")))
    except PermissionError:
        print('"负荷统计{}.xlsx"文件写入失败,请确认该文件是否已关闭！！！'.format(datetime.datetime.now().strftime("%Y-%m-%d")))
    os.system('pause')